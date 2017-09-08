from bs4 import BeautifulSoup
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from time import sleep

mainUrl = 'https://strade.sharekhan.com/rmmweb/'
dirLoc = '/path/to/location/'
NUM_STOCK = 50
NUM_ITER = 10

br = Browser()

br.open(mainUrl)
print 'Opening webpage ...'
br.select_form('login')
print 'Entering credentials ...'
br.form['loginid'] = 'username'
br.form['brpwd'] = 'br_password'
br.form['trpwd'] = 'tr_password'
resp = br.submit()
print 'Submitted credentials ...'

br.open(resp.geturl())
print 'Opening MarketWatch ...'
resp_1 = br.open('https://strade.sharekhan.com/rmmweb/mws1.sk')
print 'Reading MarketWatch page ...'

print 'Creating soup ...'
soup = BeautifulSoup(resp_1.read(), 'html.parser')
print 'Reading table ...'
table = soup.find('table', attrs = {'id' : 'sort'})

print 'Opening Excel workbook ...'
wb = Workbook()
ws = wb.active
print 'Workbook opened ...'

tableHead = []
for i in range(0, 16):
	tableHead.append(table.select('thead tr th')[i].text)

tableBody = []
flag = 1
print 'Reading table data ...'
for loop in range(0, NUM_ITER):
	t1 = (datetime.now())
	print 'Starting ', loop + 1, ' iteration at ', t1, ' ...'
	print 'Reading table data ...'
	for i in range(0, NUM_STOCK):
		for j in range(0, 16):
			tableBody.append(((table.select('tbody tr')[i]).select('td')[j].text).strip('\r\n\t;'))
	print 'Table data read ...'
	print 'Opening worksheet ...'
	for count in range(1, NUM_STOCK + 1):
		if flag == 1:
			ws.title = tableBody[(count - 1) * 16 + 3]
		else :
			ws = wb.get_sheet_by_name(tableBody[(count - 1) * 16 + 3])
		print 'Writing table headers ...'
		for i in range(0, 16):
			ws.cell(row = 1, column = (i + 1)).value = tableHead[i]
		print 'Writing table data ...'
		for j in range(0, 16):
			if j == 0 :
				ws.cell(row = 3 + loop, column = (j + 1)).value = str(datetime.now())
			else :
				ws.cell(row = 3 + loop, column = (j + 1)).value = tableBody[(count - 1) * 16 + j]

		if count != NUM_STOCK and loop == 0:
			ws = wb.create_sheet()
	
	flag = 0
	tableBody = []
	t2 = (datetime.now())
	print 'Ending ', loop + 1, ' iteration at ', t2, ' ...'
	t3 = t2 - t1
	print 'Sleeping for ', 60 - t3.total_seconds(), ' seconds ...'
	sleep(60 - t3.total_seconds())
	print 'Waking up after ', 60 - t3.total_seconds(), ' seconds ...'

print 'Saving workbook ...'
wb.save(dirLoc + 'Sharekhan_Market_Watch.xlsx')
